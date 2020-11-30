import openpyxl


class WorkBook:
    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.load_workbook(path)
        self.sheet = dict()
        self.now_sheet = None

    def change_now_sheet(self, name):
        if self.__has_sheet(name):
            self.now_sheet = self.__get_sheet(name)
        else:
            print("不存在该sheet")

        return self

    def add_sheet(self, name, index=None):
        if self.__has_sheet(name):
            print("存在该sheet")
        else:
            self.wb.create_sheet(name, index)
        return self

    def __has_sheet(self, name):
        return name in self.sheet.keys()

    def __get_sheet(self, name):
        if self.__has_sheet(name):
            return self.sheet.get(name)

    def open_sheet(self, sheet_name):
        ws = self.wb[sheet_name]
        self.sheet[sheet_name] = ws
        self.now_sheet = ws
        return self

    def change_sheet_name(self, old, new):
        if self.__has_sheet(old):
            ws = self.__get_sheet(old)
            ws.title = new
            self.sheet[new] = ws
        else:
            print("不存在该sheet")

        return self

    def __write_d1(self, wz, data):
        self.now_sheet[wz] = data

    def __write_by_sheetName1(self, name, wz, data):
        if name in self.sheet.keys():
            ws = self.sheet.get(name)
            ws[wz] = data
        else:
            print("不存在该sheet")

    def write_cell(self, wz, data, sheet_name=None):
        if sheet_name is None:
            self.__write_d1(wz, data)
        else:
            self.__write_by_sheetName1(sheet_name, wz, data)

        return self

    def __write_d2(self, row, col, data):
        ws = self.now_sheet
        ws.cell(row, col, data)

    def __write_by_sheetName2(self, name, row, col, data):
        if name in self.sheet.keys():
            ws = self.sheet.get(name)
            ws.cell(row, col, data)
        else:
            print("不存在该sheet")

    def write_by_index(self, row, col, data, sheet_name=None):
        if sheet_name is None:
            self.__write_d2(row, col, data)
        else:
            self.__write_by_sheetName2(sheet_name, row, col, data)

        return self

    def write_cells(self, data, low_row=1, high_row=1, low_col=1, high_col=1, sheet_name=None):
        if sheet_name is None:
            for row in range(low_row, high_row + 1):
                for col in range(low_col, high_col + 1):
                    self.__write_d2(row, col, data)
        else:
            for row in range(low_row, high_row + 1):
                for col in range(low_col, high_col + 1):
                    self.__write_by_sheetName2(sheet_name,row, col, data)
        return self

    def del_row(self,row_1,row_2=None,sheet_name=None):
        ws = self.now_sheet
        if sheet_name is not None:
            ws = self.__get_sheet(sheet_name)
        if row_2 is None:
            ws.delete_rows(row_1)
        else:
            for i in range(row_1,row_2):
                ws.delete_rows(i)

        return self

    def del_col(self, col_1, col_2=None, sheet_name=None):
        ws = self.now_sheet
        if sheet_name is not None:
            ws = self.__get_sheet(sheet_name)
        if col_2 is None:
            ws.delete_cols(col_1)
        else:
            for i in range(col_1, col_2+1):
                ws.delete_cols(i)
        return self

    def save(self,path=None):
        if path is None:
            self.wb.save(self.path)
        else:
            self.wb.save(path)
        return None

